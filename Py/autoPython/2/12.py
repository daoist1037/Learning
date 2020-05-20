import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)
print(wb.sheetnames[0])

sheet_1 = wb[wb.sheetnames[0]]
print(sheet_1['A1'].value)

print( 'Row '+str(sheet_1['A1'].row) + ',Column '+str(sheet_1['A1'].column) + ' is ' + sheet_1['A1'].value)